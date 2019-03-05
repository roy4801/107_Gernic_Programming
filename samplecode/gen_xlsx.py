from openpyxl import Workbook
import argparse, string

g_input_list = None
g_output_path = None
g_dry = False
def getInputPath(l):
	global g_input_list
	g_input_list = l

def getOutputPath(path):
	global g_output_path
	g_output_path = path

LABEL = {
	'Alternative': ['result/time(s)', 'input', 'Character Table Creation', 'sorting time', 'output', 'totoal', 'sort/total'],
	'Iterator': ['result/time(s)','input', 'sorting time', 'output', 'total', 'sort/total'],
	'Simple': ['result/time(s)','input', 'sorting time', 'output', 'total', 'sort/total']
}

CH = string.ascii_uppercase

def POS(i, j):
	return CH[j] + str(i+1)

def createSheetLabel(st, name):
	name = name.split('_')[0]
	for i in range(len(LABEL[name])):
		st[POS(0, i+1)] = LABEL[name][i]

MAP = {}
def getMapAlpha():
	with open('map_the_alpht.txt', 'r') as map_the_alpht_fp:
		lines = [x.split(': ') for x in map_the_alpht_fp.readlines()]
		map_tmp = []
		for i in range(18):
			map_tmp.append(lines[i][1].strip())
		MAP['line'] = map_tmp
		map_tmp = []
		for i in range(19, 37):
			map_tmp.append(lines[i][1].strip())
		MAP['len'] = map_tmp


g_file_name = {'line': 'data_line.txt', 'len': 'data_len.txt'}
g_type = ['line', 'len']
def process():
	'''
	process the data (line, len) and gen a .xlsx file
	'''
	getMapAlpha()
	wb = Workbook()
	wb.remove(wb.active)
	for i in g_input_list:
		for j in g_type:
			now = '{}_{}'.format(i, j)
			print('Creating sheet {}'.format(now))
			# if dry run, then don't do anything
			if g_dry:
				continue

			sheet = wb.create_sheet(now)
			createSheetLabel(sheet, now)

			try:
				print('>> {}'.format(i))
				with open('result/' + i + '/' + g_file_name[j], 'r') as f:
					data = [x.split() for x in f.readlines()]

					# A ~ R
					for ii in range(len(data)):
						sheet[POS(ii+1, 0)] = data[ii][0]
					# description
					for ii in range(len(MAP[j])):
						sheet[POS(ii+1, 1)] = MAP[j][ii]
					# actual data
					for ii in range(len(data)):
						for jj in range(1, len(data[ii])):
							# print('{} {}'.format(ii, jj))
							sheet[POS(ii + 1, jj-1 + 2)] = data[ii][jj]
					# total
					total_pos = 5
					percent_pos = [3, 5]
					if i == 'Alternative':
						total_pos = 6
						percent_pos = [4, 6]
					with open('result/' + i + '/data_' + j + '_total.txt', 'r') as tfp:
						total_lines = [x.split() for x in tfp.readlines()]

						for ii in range(len(total_lines)):
							sheet[POS(ii+1, total_pos)] = total_lines[ii][1]
					# sort / total
					for ii in range(len(data)):
						sheet[POS(1+ii, total_pos+1)] = '=' + POS(1+ii, percent_pos[0]) + '/' + POS(1+ii, percent_pos[1])
						sheet[POS(1+ii, total_pos+1)].style = 'Percent'
						sheet[POS(1+ii, total_pos+1)].number_format = '00.00%'

					print('=== test ===')
					for row in sheet:
						for cel in row:
							print(cel.value, end=' ')
						print()
					print('============')
			except FileNotFoundError as e:
				print('Warning: {}'.format(e))
				pass

	if not g_dry:
		print(wb.sheetnames)
		wb.save('test.xlsx')


if __name__ == '__main__':
	parser = argparse.ArgumentParser(description='Process the output time data into xlsx')
	parser.add_argument('-o', '--output', metavar='path', help='Set the output path (Default: .)')
	parser.add_argument('-i', '--input', metavar='path', help='Set the input data path (Multiple)', nargs='+')
	parser.add_argument('--dry', help='Dry run', action='store_true')
	parser.add_argument('-V', '--version', action='version', version='%(prog)s 0.0.1')

	args = parser.parse_args()

	flag = True

	if args.input:
		getInputPath(args.input)
	if args.output:
		getOutputPath(args.output)

	g_dry = args.dry
	if args.input and args.output:
		process()
		flag = False
	print('===== test ======')
	print('args.dry = {}'.format(args.dry))
	print('args.o = {}'.format(args.output))
	print('arg.i = {}'.format(args.input))
	print()
	if flag:
		parser.print_usage()